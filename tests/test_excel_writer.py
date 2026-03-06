"""Unit tests for output/excel_writer.py（spec §3.1, §3.2, AC-01, AC-03）"""
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.styles import Font

from models.book import Book
from output.excel_writer import write_excel


def _make_books() -> list[Book]:
    return [
        Book(title="好書一", price=300, isbn="9789576589065"),
        Book(title="好書二", price=450, isbn="9780000000002"),
        Book(title="好書三（長標題測試）", price=1200, isbn="9784000000003"),
    ]


class TestWriteExcel:
    def test_file_created(self):
        """AC-01：執行後確實產出 .xlsx 檔案。"""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_excel(_make_books(), output_dir=tmpdir)
            assert path.exists()
            assert path.suffix == ".xlsx"

    def test_sheet_name(self):
        """工作表名稱為「書單」（spec §3.1）。"""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_excel(_make_books(), output_dir=tmpdir)
            wb = load_workbook(path)
            assert "書單" in wb.sheetnames

    def test_header_row(self):
        """第一列標題為書名、定價、ISBN（spec §3.1）。"""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_excel(_make_books(), output_dir=tmpdir)
            ws = load_workbook(path)["書單"]
            headers = [ws.cell(1, c).value for c in range(1, 4)]
            assert headers == ["書名", "定價", "ISBN"]

    def test_header_bold(self):
        """標題列為粗體（spec §3.1）。"""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_excel(_make_books(), output_dir=tmpdir)
            ws = load_workbook(path)["書單"]
            for col in range(1, 4):
                assert ws.cell(1, col).font.bold is True

    def test_row_count(self):
        """資料列數 = 書籍數量（AC-01）。"""
        books = _make_books()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_excel(books, output_dir=tmpdir)
            ws = load_workbook(path)["書單"]
            assert ws.max_row == len(books) + 1  # +1 for header

    def test_isbn_is_string_type(self):
        """AC-03：ISBN 欄位為文字型別，不被轉為科學記號。"""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_excel(_make_books(), output_dir=tmpdir)
            ws = load_workbook(path, data_only=True)["書單"]
            for row in range(2, ws.max_row + 1):
                isbn_cell = ws.cell(row, 3)
                # 值必須是字串且為 13 碼數字
                assert isinstance(isbn_cell.value, str), (
                    f"row {row}: ISBN 應為字串，實際為 {type(isbn_cell.value)}"
                )
                assert len(isbn_cell.value) == 13
                assert isbn_cell.value.isdigit()

    def test_isbn_number_format(self):
        """ISBN 欄位的 number_format 為 '@'（文字格式）。"""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_excel(_make_books(), output_dir=tmpdir)
            ws = load_workbook(path)["書單"]
            for row in range(2, ws.max_row + 1):
                assert ws.cell(row, 3).number_format == "@"

    def test_price_is_integer(self):
        """定價欄位為整數。"""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_excel(_make_books(), output_dir=tmpdir)
            ws = load_workbook(path, data_only=True)["書單"]
            for row in range(2, ws.max_row + 1):
                price = ws.cell(row, 2).value
                assert isinstance(price, int)

    def test_empty_books(self):
        """空列表也能正常輸出（只有標題列）。"""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_excel([], output_dir=tmpdir)
            ws = load_workbook(path)["書單"]
            assert ws.max_row == 1

    def test_filename_format(self):
        """檔名格式符合 {prefix}_YYYYMMDD_HHMMSS.xlsx（spec §3.1）。"""
        import re
        with tempfile.TemporaryDirectory() as tmpdir:
            path = write_excel([], output_dir=tmpdir, prefix="suncolor_books")
            assert re.match(r"suncolor_books_\d{8}_\d{6}\.xlsx", path.name)
