"""Excel 輸出模組（spec §3.1 / §3.2）。

輸出格式：
  檔名    : suncolor_books_YYYYMMDD_HHMMSS.xlsx
  工作表  : 書單
  欄位    : ISBN | 分類 | 書名 | 定價 | 簡介 | 封面圖
  第一列  : 標題列（粗體）
  ISBN    : 強制文字格式（避免科學記號，spec AC-03）
  封面圖  : 縮圖嵌入（80×80 px）；下載失敗則留空
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Iterable

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
try:
    from PIL import Image as PILImage
except ImportError as _pil_err:
    raise ImportError(
        "Pillow 未安裝。請執行：.venv\\Scripts\\pip install Pillow\n"
        "或啟動虛擬環境後再執行 main.py。"
    ) from _pil_err

from models.book import Book

logger = logging.getLogger(__name__)

# 縮圖規格
_THUMB_W = 80   # pixels
_THUMB_H = 80   # pixels
_ROW_H_PT = 62  # points（略大於 80px，1pt ≈ 1.33px）
_COL_F_W  = 13  # 字元寬（≈ 90px）


def write_excel(
    books: Iterable[Book],
    output_dir: str = ".",
    prefix: str = "suncolor_books",
) -> Path:
    """將書籍資料寫入 Excel，回傳實際寫入的檔案路徑。"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = Path(output_dir) / f"{prefix}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "書單"

    # 標題列（粗體）
    headers = ["ISBN", "分類", "書名", "定價", "簡介", "封面圖"]
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    # 封面圖欄寬
    ws.column_dimensions["F"].width = _COL_F_W

    # 資料列（row 2 開始）
    for row_idx, book in enumerate(books, start=2):
        ws.append([
            _isbn_cell(book.isbn),  # 強制文字格式
            book.category,
            book.title,
            book.price,
            book.description,
            "",                     # 封面圖欄留空，圖片另行插入
        ])

        # ISBN 欄文字格式
        ws.cell(row=row_idx, column=1).number_format = "@"

        # 嵌入縮圖
        if book.image_url:
            xl_img = _fetch_thumbnail(book.image_url)
            if xl_img is not None:
                ws.add_image(xl_img, f"F{row_idx}")
                ws.row_dimensions[row_idx].height = _ROW_H_PT

    wb.save(filepath)
    return filepath


def _fetch_thumbnail(url: str) -> XLImage | None:
    """下載圖片並縮放為 _THUMB_W×_THUMB_H，回傳 openpyxl Image；失敗回傳 None。"""
    try:
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        pil = PILImage.open(io.BytesIO(resp.content)).convert("RGBA")
        pil.thumbnail((_THUMB_W, _THUMB_H), PILImage.LANCZOS)
        buf = io.BytesIO()
        pil.save(buf, format="PNG")
        buf.seek(0)
        xl = XLImage(buf)
        xl.width  = pil.width
        xl.height = pil.height
        return xl
    except Exception as exc:
        logger.debug("縮圖下載失敗 %s — %s", url, exc)
        return None


def _isbn_cell(isbn: str) -> str:
    """回傳字串型別的 ISBN；openpyxl 寫入字串不會轉科學記號。"""
    return str(isbn)
